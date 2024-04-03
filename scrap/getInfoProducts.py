from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import requests
import json
import translators as ts


product_counter = 0


def create_or_load_workbook():
    try:
        wb = load_workbook("products.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

        headers = [
            "Name",
            "Price",
            "In Stock",
            "Description",
            "Specifications",
            "Category",
            "Subcategory",
            "Image 1",
            "Image 2",
            "Image 3",
            "Image 4",
            "Image 5",
            "Image 6",
            "Image 7",
            "Link",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)

        wb.save("products.xlsx")

    return ws


ws = create_or_load_workbook()


def getInfoProduct(product_link, updating=False):
    global product_counter

    html = urlopen(product_link)
    bs = BeautifulSoup(html, "html.parser")
    title = bs.find(class_="detailInfo_title")
    price = bs.find(class_="shopPrice")
    detail_guide = bs.find(class_="detailGuide")
    specifications = bs.find(class_="specificationBox")

    product_counter += 1
    print(f"\nProcessing Product #{product_counter}")
    list_definitions = []
    if specifications:
        list_definitions = specifications.find_all("dl")

    categories = bs.find_all(class_="gPath_link")[2:]

    x = bs.find("script", type="application/ld+json")
    data = json.loads(
        str(x.string)
        .strip()
        .replace("&quot;", "''")
        .replace("\n", " ")
        .replace("\t", " ")
    )

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.43"
    }

    link_stock = "https://www.vevor.de/goods/goods-multi?good_sn={}".format(data["mpn"])
    response = requests.get(link_stock, headers=headers)
    qty_stock = response.json()["data"]["stockInfo"]["stock"]

    if qty_stock == "":
        qty_stock = 0

    product_name = title.text.strip()

    prodDescription = ""
    specifications_text = ""

    for list_definition in list_definitions:
        title = list_definition.find("dt")
        description = list_definition.find("dd")
        prodDescription += "{}: {};".format(
            title.text.strip(), description.text.strip()
        )

    main_category = categories[0].text.strip()
    sub_category = categories[1].text.strip()

    images = bs.find_all("img", class_="detailImg_thumbImg")
    imagesProducts = []

    for image in images:
        imagesProducts.append(image.get("src").replace("_thumb", "_img_big"))
    detail_text = ""
    if detail_guide is not None:
        detail_text = detail_guide.text

    if updating == False:
        product_name = ts.translate_text(
            query_text=product_name.replace("VEVOR", "")
            .replace("Vevor", "")
            .replace("vevor", ""),
            from_language="de",
            to_language="pt",
            translator="baidu",
        )

        if prodDescription:
            prodDescription = ts.translate_text(
                query_text=prodDescription,
                from_language="de",
                to_language="pt",
                translator="baidu",
            )

        if specifications_text:
            specifications_text = ts.translate_text(
                query_text=specifications_text,
                from_language="de",
                to_language="pt",
                translator="baidu",
            )

        if main_category:
            main_category = ts.translate_text(
                query_text=main_category,
                from_language="de",
                to_language="pt",
                translator="baidu",
            )

        if sub_category:
            sub_category = ts.translate_text(
                query_text=sub_category,
                from_language="de",
                to_language="pt",
                translator="baidu",
            )

    # Adiciona as informações diretamente na planilha
    current_row = ws.max_row + 1
    ws.cell(row=current_row, column=1, value=product_name)
    ws.cell(row=current_row, column=2, value=price["data-currency"])
    ws.cell(row=current_row, column=3, value="YES" if int(qty_stock) > 0 else "NO")
    ws.cell(row=current_row, column=4, value=prodDescription)
    ws.cell(
        row=current_row, column=5, value=specifications.text if specifications else ""
    )
    ws.cell(row=current_row, column=6, value=main_category)
    ws.cell(row=current_row, column=7, value=sub_category)

    # Adiciona as imagens nas colunas correspondentes
    for i, image_url in enumerate(imagesProducts, start=8):
        ws.cell(row=current_row, column=i, value=image_url)

    ws.cell(row=current_row, column=len(imagesProducts) + 8, value=product_link)
    # Salva a planilha após cada item raspado
    ws.parent.save("products.xlsx")

    return (
        product_name,
        price["data-currency"],
        detail_text,
        int(qty_stock) > 0,
        prodDescription,
        main_category,
        sub_category,
        imagesProducts,
        product_link,
    )
