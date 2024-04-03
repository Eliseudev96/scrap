from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import openpyxl

service = Service(ChromeDriverManager().install())

options = Options()
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
options.add_argument(f"user-agent={user_agent}")

driver = webdriver.Chrome(service=service, options=options)

workbook = openpyxl.load_workbook("products.xlsx")
sheet = workbook.active

# Verificar a última linha processada ou começar do início
last_processed_row = 2  # Inicia da linha 2 por padrão

# Iterar pelos links na planilha, processando todas as linhas a partir da última processada
for row in range(
    last_processed_row, sheet.max_row + 1
):  # Processa da última linha processada até a última linha
    try:
        product_link = sheet.cell(row, 15).value  # Os links estão na coluna "O"
        driver.get(product_link)

        # Localizando o elemento com as informações detalhadas
        detail_info_div = driver.find_element("class name", "detailInfo")

        # Extraindo informações de preço e estoque
        price_element = detail_info_div.find_element("class name", "shopPrice")
        stock_element = detail_info_div.find_element("class name", "detailStock")

        # Pegando o preço do atributo 'data-currency'
        price = price_element.get_attribute("data-currency")
        stock_status = stock_element.text

        # Atualizar a planilha com o preço e o status do estoque
        sheet.cell(row, 2).value = price  # Atualiza o preço na segunda coluna
        sheet.cell(
            row, 3
        ).value = stock_status  # Atualiza o status do estoque na terceira coluna

    except Exception as e:
        # Se ocorrer um erro, imprimir uma mensagem e continuar
        print(f"Erro na linha {row}: {str(e)}")

    finally:
        # Salvar a planilha após cada iteração para manter o progresso
        workbook.save("planilha_atualizada.xlsx")
        last_processed_row = row  # Atualiza a última linha processada

# Fechar o navegador
driver.quit()
