from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors, fills
from getInfoProducts import getInfoProduct
import translators as ts
import threading
import math

COUNTER_MAX = 0


def create_or_load_workbook():
    try:
        wb = load_workbook("products.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

        headers = [
            "Name",
            "Price",
            "In Stock",
            "Description",
            "Specifications",
            "Category",
            "Subcategory",
            "Image 1",
            "Image 2",
            "Image 3",
            "Image 4",
            "Image 5",
            "Image 6",
            "Image 7",
            "Link",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = fills.PatternFill(patternType="solid", fgColor="00BFFF")

        wb.save("products.xlsx")

    return ws


ws = create_or_load_workbook()


def split(a, n):
    k, m = divmod(len(a), n)
    return (a[i * k + min(i, m) : (i + 1) * k + min(i + 1, m)] for i in range(n))


def threadImage(ws, init, final):
    for i in range(init, final):
        try:
            if "data:image" in ws["H" + str(i)].value:
                print(i)
                product = ws["O" + str(i)].value

                (
                    name,
                    price,
                    description,
                    inStock,
                    specifications,
                    category,
                    subCategory,
                    imagesProducts,
                    product_link,
                ) = getInfoProduct(product)

                ws["A" + str(i)] = name
                ws["B" + str(i)] = price

                if inStock:
                    green = colors.Color(rgb="AAFF00")
                    fill = fills.PatternFill(patternType="solid", fgColor=green)
                    ws["C" + str(i)].fill = fill
                else:
                    red = colors.Color(rgb="EE4B2B")
                    fill = fills.PatternFill(patternType="solid", fgColor=red)
                    ws["C" + str(i)].fill = fill

                ws["C" + str(i)] = "YES" if inStock else "NO"
                ws["D" + str(i)] = description
                ws["E" + str(i)] = specifications
                ws["F" + str(i)] = category
                ws["G" + str(i)] = subCategory

                if len(imagesProducts) > 0:
                    ws["H" + str(i)] = str(imagesProducts[0])
                if len(imagesProducts) > 1:
                    ws["I" + str(i)] = str(imagesProducts[1])
                if len(imagesProducts) > 2:
                    ws["J" + str(i)] = str(imagesProducts[2])
                if len(imagesProducts) > 3:
                    ws["K" + str(i)] = str(imagesProducts[3])
                if len(imagesProducts) > 4:
                    ws["L" + str(i)] = str(imagesProducts[4])
                if len(imagesProducts) > 5:
                    ws["M" + str(i)] = str(imagesProducts[5])
                if len(imagesProducts) > 6:
                    ws["N" + str(i)] = str(imagesProducts[6])

                ws["O" + str(i)] = product_link
        except Exception as e:
            """"""
            # print(ws['I' + str(i)].value)
            print(">>", e)


def verifyImage():
    wb = load_workbook("products.xlsx")
    ws = wb.active

    count = 2

    while ws["A" + str(count)].value != None:
        count += 1

    t1 = threading.Thread(target=threadImage, args=(ws, 2, math.floor(count / 3)))
    t2 = threading.Thread(
        target=threadImage, args=(ws, math.floor(count / 3), 2 * math.floor(count / 3))
    )
    t3 = threading.Thread(
        target=threadImage, args=(ws, 2 * math.floor(count / 3), count)
    )

    t1.start()
    t2.start()
    t3.start()

    t1.join()
    t2.join()
    t3.join()

    wb.save("products.xlsx")


def threadVerify(ws, init, final):
    for i in range(init, final):
        try:
            print(i)
            if ws["A" + str(i)].value:
                product = ws["O" + str(i)].value

                (
                    name,
                    price,
                    description,
                    inStock,
                    specifications,
                    category,
                    subCategory,
                    imagesProducts,
                    product_link,
                ) = getInfoProduct(product, updating=True)

                currNameColor = ws["A" + str(i)].fill.fgColor.rgb
                if str(currNameColor) == "00FFA500" or str(currNameColor) == "FFFFA500":
                    white = colors.Color(rgb="FFFFFF")
                    fill = fills.PatternFill(patternType="solid", fgColor=white)
                    ws["A" + str(i)].fill = fill

                currColor = ws["C" + str(i)].fill.fgColor.rgb

                if (price != ws["B" + str(i)].value) or (
                    (inStock and (currColor != "00AAFF00" and currColor != "FFAAFF00"))
                    or (
                        inStock == False
                        and (currColor != "00EE4B2B" and currColor != "FFEE4B2B")
                    )
                ):
                    orange = colors.Color(rgb="FFA500")
                    fill = fills.PatternFill(patternType="solid", fgColor=orange)
                    ws["A" + str(i)].fill = fill

                    name = ts.translate_text(
                        query_text=name.replace("VEVOR", "")
                        .replace("Vevor", "")
                        .replace("vevor", ""),
                        from_language="de",
                        to_language="pt",
                        translator="baidu",
                    )

                    if description:
                        description = ts.translate_text(
                            query_text=description,
                            from_language="de",
                            to_language="pt",
                            translator="baidu",
                        )

                    print("modified: ", name)

                ws["A" + str(i)] = name
                ws["B" + str(i)] = price

                if inStock:
                    green = colors.Color(rgb="AAFF00")
                    fill = fills.PatternFill(patternType="solid", fgColor=green)
                    ws["C" + str(i)].fill = fill
                else:
                    red = colors.Color(rgb="EE4B2B")
                    fill = fills.PatternFill(patternType="solid", fgColor=red)
                    ws["C" + str(i)].fill = fill

                ws["D" + str(i)] = description
                ws["E" + str(i)] = specifications
                ws["F" + str(i)] = category
                ws["G" + str(i)] = subCategory

                if len(imagesProducts) > 0:
                    ws["H" + str(i)] = str(imagesProducts[0])
                if len(imagesProducts) > 1:
                    ws["I" + str(i)] = str(imagesProducts[1])
                if len(imagesProducts) > 2:
                    ws["J" + str(i)] = str(imagesProducts[2])
                if len(imagesProducts) > 3:
                    ws["K" + str(i)] = str(imagesProducts[3])
                if len(imagesProducts) > 4:
                    ws["L" + str(i)] = str(imagesProducts[4])
                if len(imagesProducts) > 5:
                    ws["M" + str(i)] = str(imagesProducts[5])
                if len(imagesProducts) > 6:
                    ws["N" + str(i)] = str(imagesProducts[6])

                ws["O" + str(i)] = product_link
        except Exception as e:
            print(">> ", e)
            # print(ws['I' + str(i)].value)


def verifyProducts(init=2):
    wb = load_workbook("products.xlsx")
    ws = wb.active

    count = 2

    while ws["A" + str(count)].value != None:
        count += 1

    count = 600

    # t1 = threading.Thread(target=threadVerify, args=(ws, init, math.floor(count / 3)))
    # t2 = threading.Thread(target=threadVerify, args=(ws, math.floor(count / 3), 2*math.floor(count / 3)))
    # t3 = threading.Thread(target=threadVerify, args=(ws, 2*math.floor(count / 3), count))

    # t1.start()
    # t2.start()
    # t3.start()

    # t1.join()
    # t2.join()
    # t3.join()

    # t1 = threading.Thread(target=threadVerify, args=(ws, init, count))
    # t1.start()
    # t1.joint()

    threadVerify(ws, init, count)

    wb.save("products.xlsx")


def saveInList(list_products, list_info_products):
    for product in list_products:
        try:
            p = getInfoProduct(product)
            list_info_products.append(p)
        except Exception as e:
            """"""
            # print(product)
            # print('>>', e)


def saveInSpreadsheet(products):
    wb = load_workbook("products.xlsx")
    ws = wb.active

    productsList = []

    div = len(products) // 3

    list1 = products[0:div]
    list2 = products[div : 2 * div]
    list3 = products[2 * div : len(products)]

    t1 = threading.Thread(target=saveInList, args=(list1, productsList))
    t2 = threading.Thread(target=saveInList, args=(list2, productsList))
    t3 = threading.Thread(target=saveInList, args=(list3, productsList))
    t1.start()
    t2.start()
    t3.start()

    t1.join()
    t2.join()
    t3.join()

    count = 1

    for product in productsList:
        (
            name,
            price,
            description,
            inStock,
            specifications,
            category,
            subCategory,
            imagesProducts,
            product_link,
        ) = product
        currentRow = str(count + 1)

        ws["A" + currentRow] = name
        ws["B" + currentRow] = price

        if inStock:
            green = colors.Color(rgb="AAFF00")
            fill = fills.PatternFill(patternType="solid", fgColor=green)
            ws["C" + str(currentRow)].fill = fill
        else:
            red = colors.Color(rgb="EE4B2B")
            fill = fills.PatternFill(patternType="solid", fgColor=red)
            ws["C" + str(currentRow)].fill = fill

        ws["D" + currentRow] = description
        ws["E" + currentRow] = specifications
        ws["F" + currentRow] = category
        ws["G" + currentRow] = subCategory

        if len(imagesProducts) > 0:
            ws["H" + str(currentRow)] = str(imagesProducts[0])
        if len(imagesProducts) > 1:
            ws["I" + str(currentRow)] = str(imagesProducts[1])
        if len(imagesProducts) > 2:
            ws["J" + str(currentRow)] = str(imagesProducts[2])
        if len(imagesProducts) > 3:
            ws["K" + str(currentRow)] = str(imagesProducts[3])
        if len(imagesProducts) > 4:
            ws["L" + str(currentRow)] = str(imagesProducts[4])
        if len(imagesProducts) > 5:
            ws["M" + str(currentRow)] = str(imagesProducts[5])
        if len(imagesProducts) > 6:
            ws["N" + str(currentRow)] = str(imagesProducts[6])

        ws["O" + currentRow] = product_link

        count += 1

    wb.save("products.xlsx")


def cleanTranslate():
    wb = load_workbook("products.xlsx")
    ws = wb.active
    count = 2
    while ws["A" + str(count)].value != None:
        count += 1

    for i in range(2, count):
        ws["A" + str(i)].value = (
            ws["A" + str(i)]
            .value.replace("VEVOR", "")
            .replace("Vevor", "")
            .replace("VivVor", "")
            .replace("VOR", "")
            .replace("VEFOR", "")
            .replace("Vevvor", "")
            .replace("Vorvor", "")
            .replace("vorvor", "")
            .replace("Vivvor", "")
            .replace("vivvor", "")
            .strip()
        )
        ws["C" + str(i)].value = ""

    wb.save("products.xlsx")
